import json
import openpyxl.drawing
import openpyxl.drawing.image
import requests
import openpyxl
import time
import numpy as np
from DataRetrieval import *
from collections import Counter
from openpyxl.styles import Font, Color, Fill
import matplotlib.pyplot as plt
from PIL import Image
from matplotlib.offsetbox import (OffsetImage, AnnotationBbox)


def getTopCards():
    f = open('txt/topPlayers.txt', 'r', encoding="utf-8")
    topDecks = json.loads(f.read())

    cardNames = []
    avgElixir = []
    for deck in topDecks:
        aE = 0
        for card in deck['deck']:
            cardNames.append(card['name'])
            try:
                aE += card['elixirCost']
            except:
                pass
        avgElixir.append(aE/8)

    topCards = Counter(cardNames)
    return topCards, avgElixir
    
def deckAnalysis(playerTag):
    topCards = getTopCards()[0]

    tupleTopCards = sorted(topCards.items(), key=lambda x: x[1], reverse=True)
    cardNames = [card[0] for card in tupleTopCards]
    
    playerDeckData = getPlayerData(playerTag)['currentDeck']
    playerDeck = []
    playerImages = []
    for card in playerDeckData:
        playerDeck.append(card['name'])
        playerImages.append(list(card['iconUrls'].values())[0])

    score = 0
    for card in playerDeck:
        if card in cardNames:
            score += tupleTopCards[cardNames.index(card)][1]

    p = 0
    for i in range(len(tupleTopCards)):
        p += tupleTopCards[i][1]
    
    score /= p

    fig, ax = plt.subplots()
    bars = ax.bar(cardNames, [card[1] for card in tupleTopCards])
    plt.xticks(rotation = 'vertical', fontsize = 5)
    plt.xlabel('Cards')
    plt.ylabel('Number of Uses')
    plt.title('How Many Players in the Top 1000 Use Your Cards')
    plt.subplots_adjust(bottom=0.15)

    j = 0
    for pi in playerImages:

        if playerDeck[j] in cardNames:
            x = cardNames.index(playerDeck[j])
            y = tupleTopCards[x][1]
            icon = Image.open(requests.get(pi, stream=True).raw)
            imagebox = OffsetImage(icon, zoom = 0.0375)
            ab = AnnotationBbox(imagebox, (x, y), frameon = False)
            ax.add_artist(ab)

        j += 1

    plt.savefig("plots/deckplot.png", dpi = 150)

    return playerDeckData

def elixirAnalysis(playerTag):
    XCoord = 0
    YCoord = 0
    avgElixir = getTopCards()[1]
    avgElixirDis = Counter(avgElixir)
    tupleAvgElixirDis = sorted(avgElixirDis.items(), key=lambda x: x[0])
    aeRating = ''

    playerDeckData = getPlayerData(playerTag)['currentDeck']
    ae = 0
    for card in playerDeckData:
        try:
            ae += card['elixirCost']
        except:
            pass
    
    ae /= 8

    xVals = [e[0] for e in tupleAvgElixirDis]
    yVals = [e[1] for e in tupleAvgElixirDis]

    for i in range(len(xVals)):
        if ae == xVals[i]:
            XCoord = ae-.125
            yCoord = yVals[i]
            print(ae)
            print(i)
            print(yVals[i])

    fig, ax = plt.subplots()
    bars = ax.bar(xVals, yVals, width=0.1)
    plt.xticks(xVals, rotation = 'vertical')
    plt.xlabel('Average Elixir')
    plt.ylabel('Number of Uses')
    plt.title('Your Average Elixir Among the Top 1000 Players')
    plt.subplots_adjust(bottom=0.15)
    plt.text(XCoord, yCoord, "You", fontsize = 12)

    mean = np.average(xVals,weights=yVals)
    xValsSubMean = [(x-mean)**2 for x in xVals]
    dev = np.sqrt(np.average(xValsSubMean, weights=yVals))

    if ae > mean+(dev):
        aeRating = 'Expensive Average Elixir'
    elif ae < mean-(dev):
        aeRating = 'Cheap Average Elixir'
    else:
        aeRating = 'Good Average Elixir'

    print(aeRating)
    plt.savefig("plots/elixirplot.png", dpi = 150)

    return ae

def battleLogAnalysis(playerTag):
    playerDeckData = getPlayerData(playerTag)['currentDeck']
    battleLog = getBattleLog(playerTag)
    lostCards = {}
    matches = []

    playerDeck = []
    pDN = []
    for card in playerDeckData:
        playerDeck.append(card)
        pDN.append(card['name'])

    for battle in battleLog:
        hp = 0
        OppHp = 0
        battleDeck = []
        oppDeck = []
        bDN = []
        oppName = ""
        for selctn in battle['team']:
            if selctn['princessTowersHitPoints'] is not None:
                for points in selctn['princessTowersHitPoints']:
                    hp += points
            hp += selctn['kingTowerHitPoints']
            for card in selctn['cards']:
                battleDeck.append(card)
                bDN.append(card['name'])

        for selctn in battle['opponent']:
            oppName = selctn['name']
            if selctn['princessTowersHitPoints'] is not None:
                for points in selctn['princessTowersHitPoints']:
                    OppHp += points
            OppHp += selctn['kingTowerHitPoints']
            for card in selctn['cards']:
                oppDeck.append(card)

        if sorted(bDN) == sorted(pDN):
            win = ""

            if OppHp > hp:
                win = "Loss"
                for card in oppDeck:
                    if card['name'] in lostCards.keys():
                        lostCards[card['name']][1] += 1
                    else:
                        lostCards[card['name']] = [card, 1]
            else:
                win = "Win" 

            matches.append([oppDeck,win,oppName])

    lostCards = {k: v for k, v in sorted(lostCards.items(), key=lambda item: item[1][1], reverse=True)}
    
    yVals = [v[1][1] for v in lostCards.items()]
    pi = [list(v[1][0]['iconUrls'].values())[0] for v in lostCards.items()]
    xVals = [v[0] for v in lostCards.items()]

    fig, ax = plt.subplots()
    bars = ax.bar(xVals, yVals)
    plt.xticks(rotation = 'vertical', fontsize = 5)
    plt.xlabel('Cards')
    plt.ylabel('Number of Losses')
    plt.title('How many times you lost to each card')
    plt.subplots_adjust(bottom=0.15)

    j = 0
    for p in pi:

        x = j
        y = yVals[j]
        icon = Image.open(requests.get(p, stream=True).raw)
        imagebox = OffsetImage(icon, zoom = 0.0375)
        ab = AnnotationBbox(imagebox, (x, y), frameon = False)
        ax.add_artist(ab)

        j += 1

    plt.savefig("plots/lossCardsplot.png", dpi = 150)

    return matches
    

            



def generateReport():
    playerTag = input('playerTag?')
    ##YU8LJ09R
    playerDeckData = deckAnalysis(playerTag)
    ae = elixirAnalysis(playerTag)
    matches = battleLogAnalysis(playerTag)

    playerDeck = []
    playerImages = []
    playerElixir = []
    for card in playerDeckData:
        playerDeck.append(card['name'])
        playerImages.append(list(card['iconUrls'].values())[0])
        try:
            playerElixir.append(card['elixirCost'])
        except:
            playerElixir.append(0)

    wb = openpyxl.Workbook() 
    ws = wb.active

    deckPlot = openpyxl.drawing.image.Image('plots/deckplot.png')
    elixirPlot = openpyxl.drawing.image.Image('plots/elixirplot.png')
    lossCardsPlot = openpyxl.drawing.image.Image('plots/lossCardsplot.png')
    deckPlot.anchor = ws.cell(row = 8, column = 2).coordinate
    elixirPlot.anchor = ws.cell(row = 50, column = 2).coordinate
    lossCardsPlot.anchor = ws.cell(row = 93, column = 2).coordinate
    ws.add_image(deckPlot)
    ws.add_image(elixirPlot)
    ws.add_image(lossCardsPlot)

    ws.cell(row = 6, column = 22).value = "Your Deck"
    ws.cell(row = 6, column = 22).font = Font(size = 36,bold = True)

    ws.cell(row = 6, column = 28).value = "Average Elixir: "
    ws.cell(row = 6, column = 28).font = Font(size = 24,bold = True)

    ws.cell(row = 6, column = 32).value = ae
    ws.cell(row = 6, column = 32).font = Font(size = 24)

    ws.cell(row = 25, column = 22).value = "Last 10 Matches"
    ws.cell(row = 25, column = 22).font = Font(size = 36,bold = True)

    j = 0
    for pi in playerImages:
        icon = Image.open(requests.get(pi, stream=True).raw)
        piPlot = openpyxl.drawing.image.Image(icon)
        piPlot.height = 210
        piPlot.width = 142.5
        piPlot.anchor = ws.cell(row = 8, column = 22 + (j*2)).coordinate
        ws.add_image(piPlot)

        ws.cell(row = 19, column = 22+(j*2)).value = playerDeck[j]

        j += 1
    i = 30
    m = 0
    for match in matches[::-1]:
        if m == 10:
            break
        k = 0

        ws.cell(row = i, column = 22+(k*2)).value = "Opponent: " + str(match[2])
        ws.cell(row = i+1, column = 22+(k*2)).value = match[1]
        ws.cell(row = i, column = 22).font = Font(size = 24,bold=True)
        ws.cell(row = i+1, column = 22).font = Font(size = 16)
        for card in match[0]:

            icon = Image.open(requests.get(list(card['iconUrls'].values())[0], stream=True).raw)
            miPlot = openpyxl.drawing.image.Image(icon)
            miPlot.height = 210
            miPlot.width = 142.5
            miPlot.anchor = ws.cell(row = i + 2, column = 22 + (k*2)).coordinate
            ws.add_image(miPlot)

            ws.cell(row = i + 13, column = 22+(k*2)).value = card['name']

            k += 1
        m += 1
        i += 18


    wb.save(str(time.time())+'.xlsx')



generateReport()
